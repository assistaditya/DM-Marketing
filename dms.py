import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.comments import Comment

data = pd.read_csv('anew_2 (3).csv')
rd_1 = pd.read_csv("final_lead_data_with_ratios.csv")
rd_1.columns = rd_1.columns.str.strip()

total_leads_count = data['Phone Number'].nunique()
valid_sales_calls = data[data['ConversationDuration'] > 0]
unique_phone_calls = valid_sales_calls.drop_duplicates(subset='Phone Number')
total_unique_sales_calls = len(unique_phone_calls)
phone_counts = valid_sales_calls['Phone Number'].value_counts()
repeat_follow_ups = phone_counts[phone_counts > 1]
total_follow_ups = repeat_follow_ups.count()
conversions = data[data['Delivered'] == 'Y']
total_conversions = conversions['Phone Number'].nunique()

funnel_labels = ["Leads", "Sales Call", "Follow Up", "Conversion", "Sale"]
funnel_vals = [total_leads_count, total_unique_sales_calls, total_follow_ups, total_conversions, total_conversions]
funnel_data = pd.DataFrame({"Stage": funnel_labels, "Count": funnel_vals})
funnel_fig = px.funnel(funnel_data, x='Count', y='Stage', title='Sales Funnel')

rd_1_filtered = rd_1[rd_1['Lead Month'] != 'Total']
revenue_fig = px.line(rd_1_filtered, x='Lead Month', y='Revenue_generated/Month', title='Revenue Over Time', markers=True)
lead_counts_fig = px.bar(rd_1_filtered, x='Lead Month', y='Lead_Count/Month', title='Lead Counts Over Time', text='Lead_Count/Month')
grouped_bar_fig = px.bar(rd_1, x='Lead Month', y=['Lead_Count/Month', 'Phone_Call_Count/Month', 'Validated_Lead_Count/Month', 'Count _of_Orders _Deliverd /Month'], title='Lead Counts / Phone Calls / Validated Leads / Count of Orders Delivered', barmode='group')

dual_axis_fig = go.Figure()
dual_axis_fig.add_trace(go.Bar(x=rd_1['Lead Month'], y=rd_1['Total_Cost(DM+Phone_Calls)'], name='Total Cost', marker_color='blue'))
dual_axis_fig.add_trace(go.Bar(x=rd_1['Lead Month'], y=rd_1['Revenue_generated/Month'], name='Revenue', marker_color='orange'))
dual_axis_fig.update_layout(title='Cost and Revenue by Lead Month', barmode='group', yaxis_title='Total Cost', yaxis2=dict(title='Revenue', overlaying='y', side='right'))

pie_fig = px.pie(rd_1_filtered, names='Lead Month', values='Count _of_Orders _Deliverd /Month', title='Proportion of Orders Delivered by Lead Month')

ratios = ['Cost per lead Ratio', 'Cost /confirmed_order_Month_Wise_Ratio', 'Leads to Calls Connected Ratio', 'Leads to validated Lead Ratio', 'Leads to Order Ratio', 'Roas']
ratios_figs = [px.line(rd_1_filtered, x='Lead Month', y=ratio, title=f'{ratio} Over Lead Months', markers=True) for ratio in ratios]

stacked_fig = go.Figure()
metrics = [
    ('Lead_Count/Month', 'blue', 'Lead Count'),
    ('Phone_Call_Count/Month', 'orange', 'Phone Call Count'),
    ('Validated_Lead_Count/Month', 'green', 'Validated Lead Count'),
    ('Count _of_Orders _Deliverd /Month', 'purple', 'Orders Delivered')
]

for metric, color, name in metrics:
    stacked_fig.add_trace(go.Bar(
        x=rd_1_filtered['Lead Month'], 
        y=rd_1_filtered[metric], 
        name=name, 
        marker_color=color, 
        text=rd_1_filtered[metric], 
        textposition='auto'
    ))

stacked_fig.update_layout(
    title='Monthly Metrics Overview: Leads and Orders',
    barmode='stack', 
    xaxis_title='Lead Month', 
    yaxis_title='Count', 
    legend_title='Metrics'
)

revenue_cost_fig = go.Figure()
revenue_cost_fig.add_trace(go.Bar(x=rd_1_filtered['Lead Month'], y=rd_1_filtered['Revenue_generated/Month'], name='Revenue Generated', marker_color='lightblue'))
revenue_cost_fig.add_trace(go.Bar(x=rd_1_filtered['Lead Month'], y=rd_1_filtered['Total_Cost(DM+Phone_Calls)'], name='Total Cost', marker_color='lightcoral'))
revenue_cost_fig.update_layout(title='Revenue Generated vs. Total Cost', barmode='group', xaxis_title='Lead Month', yaxis_title='Amount', legend_title='Metrics')

rd_1.columns = rd_1.columns.str.replace(' ', '_').str.replace('/', '_').str.replace('__', '_').str.strip()
filtered_rd_1 = rd_1[rd_1['Lead_Month'] != 'Total']

trend_fig = go.Figure()
trend_fig.add_trace(go.Scatter(
    x=filtered_rd_1['Lead_Month'],
    y=filtered_rd_1['Lead_Count_Month'].cumsum(),
    mode='lines+markers',
    name='Lead Count Trend',
    line=dict(color='blue', width=2, dash='dash')
))
trend_fig.add_trace(go.Scatter(
    x=filtered_rd_1['Lead_Month'],
    y=filtered_rd_1['Phone_Call_Count_Month'].cumsum(),
    mode='lines+markers',
    name='Phone Call Count Trend',
    line=dict(color='orange', width=2, dash='dash')
))
trend_fig.add_trace(go.Scatter(
    x=filtered_rd_1['Lead_Month'],
    y=filtered_rd_1['Validated_Lead_Count_Month'].cumsum(),
    mode='lines+markers',
    name='Validated Lead Count Trend',
    line=dict(color='green', width=2, dash='dash')
))
trend_fig.add_trace(go.Scatter(
    x=filtered_rd_1['Lead_Month'],
    y=filtered_rd_1['Count_of_Orders_Deliverd_Month'].cumsum(),
    mode='lines+markers',
    name='Orders Delivered Trend',
    line=dict(color='purple', width=2, dash='dash')
))

trend_fig.update_layout(
    title='Trend Lines Overview: Leads and Orders',
    xaxis_title='Month',
    yaxis_title='Count',
    legend_title='Metrics',
    template='plotly'
)

st.title("Digital Marketing Dashboard From April to September 2024")
st.write("### Lead Data Overview")
st.dataframe(rd_1)

csv = rd_1.to_csv(index=False)
st.download_button(
    label="Download Lead Data as CSV",
    data=csv,
    file_name='lead_data.csv',
    mime='text/csv'
)

excel_buffer = io.BytesIO()
with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
    rd_1.to_excel(writer, index=False, sheet_name='Lead Data')
    workbook = writer.book
    worksheet = writer.sheets['Lead Data']

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = yellow_fill

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border

    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)

    comments = {
        'Cost per lead Ratio': "Cost per Lead = Total Costs / Leads",
        'Cost /confirmed_order_Month_Wise_Ratio': "Cost per Confirmed Order = Total revenue generated in a particular month / Confirmed Orders",
        'Leads to Calls Connected Ratio': "Leads to Calls Ratio = Connected Calls / Total Leads",
        'Leads to validated Lead Ratio': "Validated Leads / Total Leads",
        'Leads to Order Ratio': "Total Order Count / Total Leads",
        'Roas': "Revenue / Advertising Spend (Total Costs, where total cost is equal to the call cost of all leads plus digital marketing costs per month.)",
        'Time to reach out to the lead Whether the person had answered the call or not is irrelevant': "Ratio (Days) = Response in Days / Time to Reach Out (Days)",
        'Time to reach out to the lead when conversation happen': "Ratio (Days) = Response in Days / Time to Reach Out (Days)",
        'Cost per validated lead': "Cost per validated Leads Ratio = Total count of validated leads / Total cost"
    }

    for col_name, comment_text in comments.items():
        if col_name in rd_1.columns:
            col_index = rd_1.columns.get_loc(col_name) + 1
            cell = worksheet.cell(row=1, column=col_index)
            cell.comment = Comment(comment_text, "System")

    notes = [
        "Cost per Lead = Total Costs / Leads",
        "Cost per Confirmed Order = Total revenue generated in a particular month / Confirmed Orders",
        "Leads to Calls Ratio = Connected Calls / Total Leads",
        "Validated Leads Ratio = Validated Leads / Total Leads",
        "Leads to Order Ratio = Orders / Total Leads",
        "Revenue / Advertising Spend (Total Costs, where total cost is equal to the call cost of all leads plus digital marketing costs per month.)",
        "",
        "Note: Validated lead data is not available for April month. (Validated data means when our telecaller team calls a lead and the lead expresses interest or requests more information.)"
    ]

    start_row = worksheet.max_row + 2
    for i, note in enumerate(notes):
        worksheet.cell(row=start_row + i, column=1, value=note)

excel_buffer.seek(0)

st.download_button(
    label="Download Lead Data as Excel",
    data=excel_buffer,
    file_name='lead_data.xlsx',
    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)

st.plotly_chart(funnel_fig)
st.write("**1. Sales Funnel:** This funnel chart illustrates the sales process from leads to conversions.")
st.write("Funnel: Leads -> Calls -> Follow-ups -> Conversions.")
st.markdown("---")

st.plotly_chart(revenue_fig)
st.write("**2. Revenue Over Time:** This plot shows the total revenue generated each month.")
st.markdown("---")

st.plotly_chart(lead_counts_fig)
st.write("**3. Lead Counts Over Time:** This bar chart represents the number of leads generated each month.")
st.markdown("---")

st.plotly_chart(grouped_bar_fig)
st.write("**4. Cost and Revenue by Month:** This chart compares total costs with revenue for each month.")
st.write("**Note:** Validated lead data is not available for April month.")
st.write("Validated data means when our telecaller team calls a lead and the lead expresses interest or requests more information.")
st.markdown("---")

st.plotly_chart(dual_axis_fig)
st.markdown("---")

st.plotly_chart(pie_fig)
st.write("**5. Proportion of Orders Delivered:** This pie chart shows the proportion of orders delivered for each month.")
st.markdown("---")

for i, (ratio_fig, description) in enumerate(zip(ratios_figs, [
    "Cost per Lead Ratio: This plot shows the cost incurred for each lead over time.",
    "Cost per Confirmed Order Ratio: This plot shows the cost incurred for each confirmed order over time.",
    "Leads to Calls Connected Ratio: This plot shows the ratio of leads that resulted in connected calls.",
    "Leads to Validated Lead Ratio: This plot shows the ratio of leads that became validated leads.",
    "Leads to Order Ratio: This plot shows the ratio of leads that resulted in orders.",
    "ROAS: Return on Ad Spend, showing revenue generated for each unit spent on ads."
])):
    st.plotly_chart(ratio_fig)
    st.write(f"**{i + 6}. {description}**")
    st.markdown("---")

st.plotly_chart(trend_fig)
st.write("**14. Trend Lines Overview:** This chart shows the trends in lead counts and orders delivered over time.")
st.markdown("---")
